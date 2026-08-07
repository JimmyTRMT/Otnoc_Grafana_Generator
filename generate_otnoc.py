
import sys
import re
from pathlib import Path

import openpyxl


# ============================================================
# Constantes fixes de l'architecture
# ============================================================
LINKED_SERVER = "INSQL"
RESOLUTION_MS = 60000
TABLE_HEIGHT_PX = 500

DEFAULT_INPUT = "OTNOC_config.xlsx"

 
# ============================================================
# 1. Lecture du fichier Excel
# ============================================================

def read_config(xlsx_path: Path):
    if not xlsx_path.exists():
        sys.exit(f"[ERREUR] Fichier introuvable : {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ----- Config -----
    if "Config" not in wb.sheetnames:
        sys.exit("[ERREUR] Onglet 'Config' manquant.")
    ws_cfg = wb["Config"]
    config = {}
    for row in ws_cfg.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        config[str(row[0]).strip()] = row[1]

    for key in ("site_name", "panel_title"):
        val = config.get(key)
        if val is None or str(val).strip() == "":
            sys.exit(f"[ERREUR] Paramètre '{key}' vide dans 'Config'.")
        if str(val).strip().lower().startswith("exemple"):
            sys.exit(
                f"[ERREUR] Le paramètre '{key}' est encore un placeholder "
                f"(« {val} »). Remplacer par la vraie valeur du site avant de générer."
            )

    # ----- OTNOC -----
    if "OTNOC" not in wb.sheetnames:
        sys.exit("[ERREUR] Onglet 'OTNOC' manquant.")
    ws = wb["OTNOC"]
    otnoc_list = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        tag, number, text = row[0], row[1], row[2]
        if tag is None and number is None and text is None:
            continue
        if not tag or number is None or not text:
            sys.exit(
                f"[ERREUR] Ligne {row_idx} de 'OTNOC' incomplète "
                f"(tag={tag!r}, number={number!r}, text={text!r})."
            )
        tag = str(tag).strip()
        text = str(text).strip()
        try:
            number = int(number)
        except (TypeError, ValueError):
            sys.exit(
                f"[ERREUR] Ligne {row_idx} : 'number' doit être un entier "
                f"(reçu {number!r})."
            )
        if any(o["tag"] == tag for o in otnoc_list):
            sys.exit(f"[ERREUR] Tag en double : {tag}")
        otnoc_list.append({"tag": tag, "number": number, "text": text})

    if not otnoc_list:
        sys.exit("[ERREUR] Aucun OTNOC dans l'onglet 'OTNOC'.")

    return config, otnoc_list


# ============================================================
# 2. Construction des 4 blocs
# ============================================================

def build_sql(otnoc_list) -> str:
    columns = ",".join(f"[{o['tag']}]" for o in otnoc_list)
    return f"""SET QUOTED_IDENTIFIER OFF

DECLARE @startTime datetime, @endTime datetime
SET @startTime = $__timeFrom()
SET @endTime = $__timeTo()


DECLARE @query nvarchar(max)
SET @query = N'
    SELECT DateTime, {columns}
    FROM WideHistory
    WHERE wwRetrievalMode = N''Delta''
    AND wwResolution = {RESOLUTION_MS}
    AND wwQualityRule = N''Extended''
    AND wwVersion = N''Latest''
    AND wwTimeZone = N''UTC''
   
    AND DateTime >= ''' + CONVERT(nvarchar(23), @startTime, 121) + N'''
    AND DateTime <= ''' + CONVERT(nvarchar(23), @endTime, 121) + N'''
    ORDER BY DateTime DESC
'

DECLARE @openQuery nvarchar(max)
SET @openQuery = N'
    SELECT *
    FROM OPENQUERY({LINKED_SERVER}, N''' + REPLACE(@query, '''', '''''') + N''')
'

EXEC(@openQuery)
"""


def build_sql_pivot(otnoc_list) -> str:
    """Variante sans WideHistory.

    Les tags ne sont plus des noms de colonnes envoyés au provider INSQL mais
    de simples chaînes dans un IN(...) sur la table History ; le pivot est fait
    par SQL Server. Indispensable quand les tags contiennent des points
    (nomenclature type AU_MEM_F1.AU1.PVFL), que le provider n'arrive pas à
    résoudre en colonnes de WideHistory.
    """
    in_list = ",".join(f"''{o['tag']}''" for o in otnoc_list)
    columns = ",".join(f"[{o['tag']}]" for o in otnoc_list)
    return f"""SET QUOTED_IDENTIFIER OFF

DECLARE @startTime datetime, @endTime datetime
SET @startTime = $__timeFrom()
SET @endTime = $__timeTo()


DECLARE @query nvarchar(max)
SET @query = N'
    SELECT DateTime, TagName, Value
    FROM History
    WHERE TagName IN ({in_list})
    AND wwRetrievalMode = N''Delta''
    AND wwResolution = {RESOLUTION_MS}
    AND wwQualityRule = N''Extended''
    AND wwVersion = N''Latest''
    AND wwTimeZone = N''UTC''

    AND DateTime >= ''' + CONVERT(nvarchar(23), @startTime, 121) + N'''
    AND DateTime <= ''' + CONVERT(nvarchar(23), @endTime, 121) + N'''
'

DECLARE @openQuery nvarchar(max)
SET @openQuery = N'
    SELECT DateTime, {columns}
    FROM (
        SELECT DateTime, TagName, Value
        FROM OPENQUERY({LINKED_SERVER}, N''' + REPLACE(@query, '''', '''''') + N''')
    ) AS src
    PIVOT (MAX(Value) FOR TagName IN ({columns})) AS pvt
    ORDER BY DateTime DESC
'

EXEC(@openQuery)
"""


def build_html() -> str:
    return """<div class="tableFixHead">
  <button id="SaveAs">Exporter en CSV</button>
  <table id="tableau">
    <thead>
      <tr id="header">
        <th>Date de début</th>
        <th>Heure de début</th>
        <th>Date de fin</th>
        <th>Heure de fin</th>
        <th>Temps</th>
        <th>Défaillance</th>
        <th>n° OTNOC</th>
      </tr>
    </thead>
    <tbody id="content">
      </tbody>
  </table>
</div>
"""


def build_css() -> str:
    return f""".tableFixHead {{ 
  overflow-y: auto; 
  height: {TABLE_HEIGHT_PX}px; 
  background-color: #000000 !important; 
}}

#tableau {{
  border-collapse: collapse;
  width: 100%;
  background-color: #000000 !important;
  font-family: sans-serif;
}}

#tableau td {{
  padding: 10px;
  border: 1px solid #333333;
  color: #ffffff !important;
  background-color: #000000 !important;
  font-size: 13px;
  text-align: left;
}}

#tableau th {{
  background-color: #222222 !important;
  color: #ffffff !important;
  font-weight: bold;
  padding: 12px;
  position: sticky;
  top: 0;
  border: 1px solid #444444;
  text-transform: uppercase;
  font-size: 12px;
  z-index: 2;
  text-align: left;
}}

#tableau tr {{
  background-color: #000000 !important;
}}

#tableau tr:hover td {{
  background-color: #1a1a1a !important;
}}

#SaveAs {{
  background-color: #333333;
  color: #ffffff;
  border: 1px solid #555555;
  padding: 8px 15px;
  margin-bottom: 10px;
  cursor: pointer;
  border-radius: 4px;
}}

#SaveAs:hover {{
  background-color: #444444;
}}
"""


DETECTION_BOOL = """function calculateStateChanges(dateArray, valueArray) {
    const stateChanges = [];
    if (valueArray.length === 0) return stateChanges;

    var currentState = valueArray[0];
    var startDate = dateArray[0];

    for (var i = 1; i < valueArray.length; i++) {
        const currentDate = dateArray[i];
        const currentValue = valueArray[i];
        if (currentValue === null) continue;

        if (currentValue !== currentState) {
            if (currentDate > startDate) {
                stateChanges.push({
                    state: currentState,
                    startDate: startDate,
                    endDate: currentDate
                });
            }
            currentState = currentValue;
            startDate = currentDate;
        }
    }
    return stateChanges;
}

var final_array = [];
var dates = valueField[0].values;

for (var y = 1; y < valueField_size; y++) {
    const stateData = valueField[y].values;
    const tagName = valueField[y].name;
    const stateDurations = calculateStateChanges(dates, stateData);

    stateDurations.forEach(d => {
        if (d.state === 1) {
            final_array.push([d.startDate, d.endDate, tagName]);
        }
    });
}

"""

DETECTION_INCREMENT = """function calculateIncrements(dateArray, valueArray) {
    const increments = [];
    var previousValue = null;
    var startDate = null;
    var lastDate = null;

    for (var i = 0; i < valueArray.length; i++) {
        const currentValue = valueArray[i];
        if (currentValue === null) continue;

        const value = Number(currentValue);
        if (previousValue === null) {
            previousValue = value;
            continue;
        }

        if (value > previousValue) {
            if (startDate === null) startDate = dateArray[i];
            lastDate = dateArray[i];
        } else if (startDate !== null) {
            increments.push({ startDate: startDate, endDate: lastDate });
            startDate = null;
        }
        previousValue = value;
    }

    if (startDate !== null) increments.push({ startDate: startDate, endDate: lastDate });
    return increments;
}

var final_array = [];
var dates = valueField[0].values;

for (var y = 1; y < valueField_size; y++) {
    const counterData = valueField[y].values;
    const tagName = valueField[y].name;
    const occurrences = calculateIncrements(dates, counterData);

    occurrences.forEach(d => {
        final_array.push([d.startDate, d.endDate, tagName]);
    });
}

"""


def build_onrender(site_name: str, otnoc_list, mode: str) -> str:
    """Bloc onRender. Un mode = un bloc, avec sa seule logique de détection."""
    dict_lines = []
    for o in otnoc_list:
        text_escaped = o["text"].replace("\\", "\\\\").replace('"', '\\"')
        dict_lines.append(
            f'    "{o["tag"]}": {{ number: {o["number"]}, '
            f'text: "{text_escaped}" }},'
        )
    dict_body = "\n".join(dict_lines)
    detection = DETECTION_INCREMENT if mode == "increment" else DETECTION_BOOL

    return f"""console.clear();

function messageTableau(msg) {{
    var cible = htmlNode.getElementById("content");
    if (cible) cible.innerHTML = `<tr><td colspan="7">${{msg}}</td></tr>`;
}}

if (!data.series || data.series.length === 0) {{
    messageTableau("Aucune série renvoyée par Query A verifié le Tag OTNOC.");
    return;
}}

var valueField = data.series[0].fields;
var valueField_size = valueField.length;

if (valueField_size < 2) {{
    messageTableau("Query A ne renvoie aucune donnée sur la période (séries : " + data.series.length
        + ", colonnes : " + valueField_size + "). Les tags existent mais n'ont aucune valeur historisée sur cette fenêtre de temps.");
    return;
}}

if (!valueField[0].values || valueField[0].values.length === 0) {{
    messageTableau("Query A renvoie 0 ligne sur la période sélectionnée.");
    return;
}}

const defaut_data = {{
{dict_body}
}};

function toHHMMSS(secs) {{
    var sec_num = parseInt(secs, 10);
    var hours = String(Math.floor(sec_num / 3600)).padStart(2, "0");
    var minutes = String(Math.floor(sec_num / 60) % 60).padStart(2, "0");
    var seconds = String(sec_num % 60).padStart(2, "0");
    return hours + ":" + minutes + ":" + seconds;
}}

function timeConverter(t) {{
    var a = new Date(t);
    var year = a.getUTCFullYear();
    var month_nb = String(a.getUTCMonth() + 1).padStart(2, "0");
    var date = String(a.getUTCDate()).padStart(2, "0");
    var hour = String(a.getUTCHours()).padStart(2, "0");
    var min = String(a.getUTCMinutes()).padStart(2, "0");
    var sec = String(a.getUTCSeconds()).padStart(2, "0");

    var date_date = date + "/" + month_nb + "/" + year;
    var date_hour = hour + ":" + min + ":" + sec;

    return [date_date, "", date_hour];
}}

{detection}
final_array.sort((a, b) => b[0] - a[0]);

var tableau = htmlNode.getElementById("content");
var tableau_html = [];
var values_csv = [];

final_array.forEach(item => {{
    var startInfo = timeConverter(item[0]);
    var endInfo = timeConverter(item[1]);
    var tempsSec = (item[1] - item[0]) / 1000;
    var tempsStr = toHHMMSS(tempsSec);

    var tag_name = item[2];
    var tag_info = defaut_data[tag_name] || {{ number: "??", text: "Tag Inconnu (" + tag_name + ")" }};

    tableau_html.push(`<tr>
        <td>${{startInfo[0]}}</td><td>${{startInfo[2]}}</td>
        <td>${{endInfo[0]}}</td><td>${{endInfo[2]}}</td>
        <td>${{tempsStr}}</td><td>${{tag_info.text}}</td><td>${{tag_info.number}}</td>
    </tr>`);

    values_csv.push([startInfo[0], startInfo[2], endInfo[0], endInfo[2], tempsStr, tag_info.text, tag_info.number]);
}});

tableau.innerHTML = tableau_html.join("");

var tableau_header_array = ["Date de début", "Heure de début", "Date de fin", "Heure de fin", "Temps", "Défaillance", "n° OTNOC"];
var export_csv = tableau_header_array.join(";") + "\\r\\n";
values_csv.forEach(row => {{
    export_csv += row.join(";") + "\\r\\n";
}});

function downloadBlob(content, filename, contentType) {{
    const BOM = "\\uFEFF";
    var blob = new Blob([BOM + content], {{ type: contentType }});
    var url = URL.createObjectURL(blob);
    var pom = document.createElement("a");
    pom.href = url;
    pom.setAttribute("download", filename);
    pom.click();
}}

var exportCSVBtn = htmlNode.getElementById("SaveAs");
if(exportCSVBtn) {{
    exportCSVBtn.onclick = () => downloadBlob(export_csv, "Export_OTNOC_{site_name}.csv", "text/csv;charset=utf-8;");
}}
"""


# ============================================================
# 3. Assemblage du .txt final
# ============================================================

def section(title: str) -> str:
    bar = "=" * 72
    return f"\n\n{bar}\n{title}\n{bar}\n\n"


def build_txt(config, otnoc_list) -> str:
    site_name = str(config["site_name"]).strip()
    panel_title = str(config["panel_title"]).strip()

    header = (
        f"OTNOC - \n"
        f"Site         : {site_name}\n"
        f"Titre panel  : {panel_title}\n"
        f"Nb d'OTNOC   : {len(otnoc_list)}\n"
        f"Linked Server: {LINKED_SERVER}   |   Résolution: {RESOLUTION_MS} ms   "
        f"|   Hauteur: {TABLE_HEIGHT_PX} px\n"
        f"\n"
        f"Coller chaque bloc ci-dessous dans la zone correspondante du panel\n"
        f"Grafana. Aucun autre réglage requis.\n"
        f"\n"
        f"ATTENTION : il y a DEUX blocs onRender, un par codage de la valeur.\n"
        f"N'en coller qu'UN SEUL, celui qui correspond au site :\n"
        f"  - POUR LE BOOLEEN       : 0 = pas de défaut, 1 = défaut actif.\n"
        f"  - POUR L'INCREMENTATION : la valeur est un compteur qui avance à\n"
        f"                            chaque occurrence.\n"
    )

    parts = [
        header,
        section("À mettre dans Query A (data source SQL du site)"),
        build_sql(otnoc_list),
        section(
            "VARIANTE de Query A - à utiliser SEULEMENT si la requête ci-dessus\n"
            "renvoie « error occurred while preparing the query ... INSQL »\n"
            "(cas des tags contenant des points). Même résultat, sans WideHistory."
        ),
        build_sql_pivot(otnoc_list),
        section("À mettre dans HTML/SVG document"),
        build_html(),
        section("À mettre dans CSS"),
        build_css(),
        section(
            "À mettre dans onRender - VERSION 1/2 : POUR LE BOOLEEN\n"
            "0 = pas de défaut, 1 = défaut actif.\n"
            "Ne coller que ce bloc OU le suivant, jamais les deux."
        ),
        build_onrender(site_name, otnoc_list, "bool"),
        section(
            "À mettre dans onRender - VERSION 2/2 : POUR L'INCREMENTATION\n"
            "La valeur est un compteur qui avance à chaque occurrence.\n"
            "Ne coller que ce bloc OU le précédent, jamais les deux."
        ),
        build_onrender(site_name, otnoc_list, "increment"),
    ]
    return "".join(parts)


# ============================================================
# 4. Main
# ============================================================

def slug(text: str) -> str:
    text = re.sub(r"[^\w\-]+", "_", text.strip())
    return text.strip("_") or "site"


def main():
    xlsx_path = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT)

    config, otnoc_list = read_config(xlsx_path)
    txt_content = build_txt(config, otnoc_list)

    site = slug(str(config["site_name"]))
    out_path = Path(f"OTNOC_{site}.txt")
    out_path.write_text(txt_content, encoding="utf-8")

    print(f"OK - Fichier généré : {out_path.resolve()}")
    print(f"     Site           : {config['site_name']}")
    print(f"     Panel          : {config['panel_title']}")
    print(f"     OTNOC          : {len(otnoc_list)} défaut(s)")
    print(f"     Numéros        : {sorted(o['number'] for o in otnoc_list)}")


if __name__ == "__main__":
    main()
